import os
import json
import openpyxl

def excel_to_json(excel_path, json_path, sheet_name="Hotels"):
    print(f"Reading Excel file: {excel_path} (Sheet: {sheet_name})")
    if not os.path.exists(excel_path):
        print(f"Error: Excel file '{excel_path}' does not exist!")
        return

    try:
        # Load workbook and sheet
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in {excel_path}!")
            return
            
        ws = wb[sheet_name]
        
        # Read rows
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("Warning: The sheet is empty!")
            return
            
        # First row is the header
        headers = rows[0]
        data = []
        
        for row in rows[1:]:
            # If the row is completely empty, skip it
            if all(val is None for val in row):
                continue
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if header is not None:
                    row_dict[header] = row[col_idx]
            data.append(row_dict)
            
        # Show first 5 records
        print("\nFirst 5 records:")
        for record in data[:5]:
            print(json.dumps(record, indent=2))
            
        # Print schema-like header information
        print("\nColumns / Schema:")
        for h in headers:
            if h is not None:
                print(f" - {h}")
                
        # Write to JSON file
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
            
        print(f"\nSuccessfully converted! JSON saved at: {json_path}")
        
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    excel_to_json("chennai.xlsx", "./results/chennai.json")